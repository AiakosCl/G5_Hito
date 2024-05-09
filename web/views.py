from django.db.models.base import Model as Model
from django.db.models.query import QuerySet
from django.shortcuts import render, redirect, get_object_or_404
from .forms import *
from .models import *
from django.urls import reverse_lazy
from django.contrib import messages
from django.contrib.auth.views import LoginView, LogoutView
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.db.models import Q #Para generar búsqdas en la base de datos.
import openpyxl #Para trabajar con archivos Excel

# Create your views here.

iconos = {
    'ok':'👌🏼',
    'mal':'👎🏼'
}

# Vista para index.html
def index(request):
    productos = Productos.objects.filter(privado=False)
    return render(request, 'index.html',{'productos':productos})

# vista para about.html
def about(request):
    return render(request, 'about.html',{})

# Vista para welcome.html
@login_required
def welcome(request):
    productos = Productos.objects.filter(privado=True)
    return render(request, 'welcome.html',{'productos':productos})


# --- TRABAJO CON MODELO PRODUCTO ------
# Vista para crear un nuevo producto
@login_required
def nuevo_producto(request):
    if request.method == 'POST':
        formulario = ProductoForm(request.POST, request.FILES)
        if formulario.is_valid():
            producto = formulario.save(commit=False)
            producto.save()
            messages.success(request, f'{iconos["ok"]}\tEl producto ha sido creado con éxito!')
            return redirect('DetalleProducto', slug=producto.slug)
        else:
            messages.warning(request,f'{iconos["mal"]}\tPor favor, revisar la información faltante.')
    else:
        formulario = ProductoForm()
    return render(request, 'producto.html', {'formulario':formulario})

# Vista para ver el registro creado o modificado.
@login_required
def detalle_producto(request, slug):
    producto = get_object_or_404(Productos, slug=slug)
    return render(request, 'detalle_producto.html', {'producto':producto})

# Vista para ver todos los productos creados, y que permitirá editarlos o borrarlos.
@login_required
def lista_productos(request):
    productos = Productos.objects.all()
    return render(request, 'lista_productos.html',{'producto':productos})

# Vista para ver más detalles del un producto en específico.
def vista_producto(request, producto_id):
    flan = get_object_or_404(Productos, pk = producto_id)
    return render(request, 'info.html', {'producto':flan})

# Vista para generar una lista filtrada por un criterio en el nombre o en la descripción
def filtrar(request):
    termino = request.GET.get('q')
    productos = Productos.objects.all()

    if not request.user.is_authenticated:
        productos = productos.filter(privado=False)
    
    if termino:
        productos = productos.filter(
            Q(nombre__icontains=termino)|Q(descripcion__icontains=termino)
        )
    
    return render(request, 'index.html', {'productos':productos, 'seccion':'Menu'})

# Vista para eliminar un producto en específico.
@login_required
def eliminar_producto(request, producto_id):
    if request.user.is_superuser or request.user.is_staff:    
        try:
            producto = Productos.objects.get(pk=producto_id)
        except Productos.DoesNotExist:
            return "Producto no se encuenta"
        
        if request.method == 'POST':
            producto.delete()
            messages.success(request, f'{iconos["ok"]}\tSe ha eliminado el producto.')
            return redirect('ListaProducto')
        
        return render(request, 'producto_eliminar.html', {'producto':producto})
    else:
        messages.warning(request,f'{iconos["mal"]}\tUsted no está autorizado para esta operación')
        return redirect('Inicio')

# Vista para editar un producto en específico.
@login_required
def editar_producto(request, producto_id):
    if request.user.is_superuser or request.user.is_staff:    
        try:
            producto = Productos.objects.get(pk=producto_id)
        except Productos.DoesNotExist:
            return "Producto no se encuenta"
        
        if request.method == 'POST':
            producto.nombre = request.POST['nombre']
            producto.descripcion = request.POST['descripcion']
            
            if 'imagen' in request.FILES:
                producto.imagen = request.FILES['imagen']
            
            producto.precio = request.POST['precio']
            producto.liviano = request.POST['liviano']
            producto.privado = request.POST['privado']
            producto.save()
            messages.success(request, f'{iconos["ok"]}\tSe ha modificado el producto.')
            return redirect('ListaProducto')
        
        return render(request, 'producto_editar.html', {'formulario':producto})
    else:
        messages.warning(request,f'{iconos["mal"]}\tUsted no está autorizado para esta operación')
        return redirect('Inicio')

# Vista para cargar de forma masiva la tabla de productos
@login_required
def carga_masiva(request):
    if request.method == 'POST' and request.FILES['archivo_excel']:
        archivo_excel = request.FILES['archivo_excel']
        try:
            wb = openpyxl.load_workbook(archivo_excel) #Cargará en wb el Workbook de Excel
            ws = wb.active #Cargará en ws el Worksheet activo para trabajar

            encabezados = [cell.value for cell in ws[1]] #Se está haciendo una lista con los encabezados.
            campos_requeridos = ['nombre','descripcion'] #Se definen cuáles son los campos requeridos (según nuestro modelo)


            for row in ws.iter_rows(min_row=2, values_only=True):
                datos_producto=dict(zip(encabezados, row))
                # nombre, descripcion, imagen, privado = row

                for campo in campos_requeridos:
                    if campo not in datos_producto or not datos_producto[campo]:
                        raise ValueError(f"El campo '{campo}' es requerido.") #Se está validando si el valor de nombre o descripción no sea nulo.
                
                nombre = datos_producto['nombre'].strip() #Se está removiendo los espacios al principio y final de valor de la celda para un carga más limpia.
                descripcion = datos_producto['descripcion'].strip()
                imagen = 'productos/cargar_imagen.png' if datos_producto.get('imagen') is None else datos_producto['imagen'] #La función Get obtendrá el valor del campo, si no hay respuesta, entregará un valor por default
                precio = 0 if datos_producto.get('precio') is None else float(datos_producto['precio'])
                liviano = bool(datos_producto.get('liviano', False))
                privado = bool(datos_producto.get('privado', True))

                #Acá irá creando los registros en el Modelo Productos
                Productos.objects.create(
                    nombre=nombre,
                    descripcion = descripcion,
                    imagen = imagen,
                    privado = privado
                )
            
            messages.success(request,f"{iconos['ok']}\t Toda la información se cargó correctamente!!!")
            return redirect('ListaProducto')

        except FileNotFoundError:
            messages.error(request, f'{iconos["mal"]}\tNo se encontró el archivo Excel.')
        except (KeyError, ValueError) as e:
            messages.error(request, f'{iconos["mal"]}\tError en el formato del archivo Excel: {str(e)}')
        except Exception as e:
            messages.error(request, f'{iconos["mal"]}\tOcurrió un error durante la carga de productos: {str(e)}')
    
    return render(request, 'carga_masiva.html')


# ---- TRABAJO CON MODELO CONTACTO ---------
# Vista para ver la Rutina de contacto.
def contact_view(request):
    if request.method == 'POST':
        formulario = ContactoForm(request.POST)
        if formulario.is_valid():
            formulario.save()
            messages.success(request, f'{iconos["ok"]}\t¡Gracias por su contacto!')
            return redirect('Inicio')
        else:
            messages.warning(request,f'{iconos["mal"]}\tPor favor, revisar la información ingresada.')
    else:
        formulario = ContactoForm()
    return render(request, 'contacto.html', {'contacto':formulario})



# Uso de Clases para Login y Logout
# Vista Login y Logout de Django

class Login(LoginView):
    template_name="registration/login.html"

class Logout(LogoutView):
    next_page="/"


# --- TRABAJO CON EL MODELO USER ---
# vistas para trabajo con usuarios (Creación, edición, eliminación)

@login_required
def nuevo_usuario(request):
    if request.method=='POST':
        formulario = NuevoUsuarioForm(request.POST)
        print(formulario)
        if formulario.is_valid():
            try:
                usuario = formulario.save(commit=False)
                usuario.save()
                # login(request, usuario) #Para iniciar sesión automáticamente, si se desea.
                messages.success(request, f'{iconos["ok"]}\t¡Se ha registrado al nuevo usuario!')
                return redirect('Ficha', usuario_id=usuario.id)
            except ValueError as e:
                messages.error(request, f'{iconos["mal"]}\tError al crear el usuario: {str(e)}')
        else:
            messages.warning(request,f'{iconos["mal"]}\tUps! Algo salió mal. Revisar la inforamción ingresada.')
    else:
        formulario = NuevoUsuarioForm()

    return render(request, 'usuario_nuevo.html', {'formulario':formulario})

@login_required
def lista_usuarios(request):
    usuarios = User.objects.all()
    return render(request, 'lista_usuarios.html',{'lista':usuarios})

@login_required
def vista_ficha(request, usuario_id):
    usuario = get_object_or_404(User, pk=usuario_id)
    return render(request, 'detalle_usuario.html', {'usuario':usuario})

@login_required
def editar_usuario(request, usuario_id):
    if request.user.is_superuser or request.user.is_staff:    
        try:
            usuario = User.objects.get(pk=usuario_id)
        except User.DoesNotExist:
            return "Usuario no Existe"
        
        if request.method == 'POST':
            usuario.first_name = request.POST['first_name']
            usuario.last_name = request.POST['last_name']
            usuario.email = request.POST['email']
            usuario.is_staff = request.POST['is_staff']
            usuario.is_active = request.POST['is_active']
            
            if request.POST['password']:
                usuario.set_password(request.POST['password']) #set_password establece una password encriptada para la tabla User

            usuario.save()
            messages.success(request, f'{iconos["ok"]}\t¡Actualización de datos correcta!')
            return redirect('ListaUsuarios')
        
        return render(request, 'usuario_editar.html', {'formulario':usuario})
    else:
        messages.warning(request,f'{iconos["mal"]}\tUsted no está autorizado para esta operación')
        return redirect('Inicio')

@login_required
def eliminar_usuario(request, usuario_id):
    if request.user.is_superuser or request.user.is_staff:    
        try:
            usuario = User.objects.get(pk=usuario_id)
        except User.DoesNotExist:
            return "Usuario no existe"
        
        if request.method == 'POST':
            usuario.delete()
            messages.success(request, f'{iconos["ok"]}\t¡Se ha eliminado el usuario!')
            return redirect('ListaUsuarios')
        
        return render(request, 'usuario_eliminar.html', {'usuario':usuario})
    else:
        messages.warning(request,f'{iconos["mal"]}\tUsted no está autorizado para esta operación')
        return redirect('Inicio')